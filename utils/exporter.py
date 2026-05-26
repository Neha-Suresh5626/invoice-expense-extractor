"""
utils/exporter.py
User-specific professional Excel export
"""

import os
import pandas as pd

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)


def export_to_excel(
    data: dict,
    output_folder: str,
    username: str
) -> str:

    # -----------------------------------
    # CREATE OUTPUT FOLDER
    # -----------------------------------

    os.makedirs(output_folder, exist_ok=True)

    # -----------------------------------
    # USER-SPECIFIC FILE
    # -----------------------------------

    filename = f"{username}_invoices.xlsx"

    filepath = os.path.join(
        output_folder,
        filename
    )

    # -----------------------------------
    # NEW ROW DATA
    # -----------------------------------

    row = {

        "Vendor":
            data.get("vendor_name", "Not Available"),

        "Invoice No.":
            data.get("invoice_number", "Not Available"),

        "Date":
            data.get("date", "Not Available"),

        "Due Date":
            data.get("due_date", "Not Available"),

        "Category":
            data.get("category", "General"),

        "Currency":
            data.get("currency", "INR"),

        "Subtotal":
            data.get("subtotal", "0"),

        "Tax":
            data.get("tax", "0"),

        "GST":
            data.get("gst", "0"),

        "Total":
            data.get("total", "0"),

        "Type":
            data.get("invoice_type", "typed"),

        "Processed":
            datetime.now().strftime(
                "%d %b %Y %H:%M"
            )
    }

    df_new = pd.DataFrame([row])

    # -----------------------------------
    # APPEND EXISTING DATA
    # -----------------------------------

    if os.path.exists(filepath):

        df_old = pd.read_excel(filepath)

        df_final = pd.concat(
            [df_old, df_new],
            ignore_index=True
        )

    else:

        df_final = df_new

    # -----------------------------------
    # SAVE TO EXCEL
    # -----------------------------------

    df_final.to_excel(
        filepath,
        index=False,
        sheet_name="Invoices"
    )

    # -----------------------------------
    # LOAD WORKBOOK
    # -----------------------------------

    wb = load_workbook(filepath)

    ws = wb.active

    # -----------------------------------
    # COLORS / STYLES
    # -----------------------------------

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="0A0C0F"
    )

    ROW_FILL_1 = PatternFill(
        fill_type="solid",
        fgColor="111418"
    )

    ROW_FILL_2 = PatternFill(
        fill_type="solid",
        fgColor="1A1F27"
    )

    THIN = Side(
        style="thin",
        color="2A3040"
    )

    BORDER = Border(
        left=THIN,
        right=THIN,
        top=THIN,
        bottom=THIN
    )

    # -----------------------------------
    # HEADER STYLING
    # -----------------------------------

    for cell in ws[1]:

        cell.font = Font(
            name="Calibri",
            bold=True,
            color="00E5A0",
            size=11
        )

        cell.fill = HEADER_FILL

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = BORDER

    # -----------------------------------
    # DATA ROW STYLING
    # -----------------------------------

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2),
        start=2
    ):

        fill = (
            ROW_FILL_1
            if row_idx % 2 == 0
            else ROW_FILL_2
        )

        for cell in row:

            cell.font = Font(
                name="Calibri",
                size=10,
                color="FFFFFF"
            )

            cell.fill = fill

            cell.border = BORDER

            cell.alignment = Alignment(
                vertical="center"
            )

    # -----------------------------------
    # AUTO COLUMN WIDTH
    # -----------------------------------

    for col in ws.columns:

        max_len = 0

        column_letter = col[0].column_letter

        for cell in col:

            try:

                value = str(cell.value)

                if len(value) > max_len:

                    max_len = len(value)

            except:
                pass

        adjusted_width = min(max_len + 4, 45)

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

    # -----------------------------------
    # FREEZE HEADER ROW
    # -----------------------------------

    ws.freeze_panes = "A2"

    # -----------------------------------
    # SAVE FINAL FILE
    # -----------------------------------

    wb.save(filepath)

    return filepath